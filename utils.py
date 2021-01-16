import csv
from os.path import isfile
from typing import List

import yaml
from openpyxl import load_workbook, Workbook

from options import VesselDetectorOptions
from results import VesselDetectorResult


def write_results(results: List[VesselDetectorResult], options: VesselDetectorOptions, stem: str):
    # YAML
    yaml_path = f"{stem}.results.yml"
    print(f"Writing YAML file: {yaml_path}")
    with open(yaml_path, 'w') as file:
        yaml.dump(results, file, default_flow_style=False)

    # CSV
    csv_path = f"{stem}.results.csv"
    print(f"Writing CSV file: {csv_path}")
    with open(csv_path, 'w') as file:
        writer = csv.writer(file, delimiter=',', quotechar='|', quoting=csv.QUOTE_MINIMAL)
        writer.writerow(['ID', 'Area', 'Solidity', 'Max Width', 'Max Height'])
        for result in results:
            writer.writerow([result.id, result.area, result.solidity, result.max_height, result.max_width])

    # Excel
    excel_path = f"{stem}.results.xlsx"
    print(f"Writing Excel file: {excel_path}")
    wb = load_workbook(excel_path) if isfile(excel_path) else Workbook()
    sheet = wb.active
    sheet.cell(row=1, column=1).value = 'filename'
    sheet.cell(row=1, column=2).value = 'leaf_area'
    sheet.cell(row=1, column=3).value = 'solidity'
    sheet.cell(row=1, column=4).value = 'max_width'
    sheet.cell(row=1, column=5).value = 'max_height'
    for result in results:
        sheet.append([result.id, result.area, result.solidity, result.max_height, result.max_width])
    wb.save(excel_path)