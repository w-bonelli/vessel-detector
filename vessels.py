import csv
from os.path import join, isfile
from pathlib import Path

import click
import yaml
from openpyxl import load_workbook, Workbook

from options import VesselDetectorOptions
from traits import extract_traits


@click.command()
@click.argument('input_file')
@click.option('-o', '--output_directory', required=False, type=str, default='')
@click.option('-mr', '--min_radius', required=False, type=int, default=15)
@click.option('-ms', '--min_size', required=False, type=int, default=500)
def cli(input_file, output_directory, min_radius, min_size):
    Path(output_directory).mkdir(parents=True, exist_ok=True)
    options = VesselDetectorOptions(
        input_file=input_file,
        output_directory=output_directory,
        min_radius=min_radius,
        min_cluster_size=min_size)
    results = extract_traits(options)

    # YAML
    yaml_path = join(output_directory, f"{options.input_stem}.results.yml")
    print(f"Writing YAML file: {yaml_path}")
    with open(yaml_path, 'w') as file:
        yaml.dump(results, file, default_flow_style=False)

    # CSV
    csv_path = join(output_directory, f"{options.input_stem}.results.csv")
    print(f"Writing CSV file: {csv_path}")
    with open(csv_path, 'w') as file:
        writer = csv.writer(file, delimiter=',', quotechar='|', quoting=csv.QUOTE_MINIMAL)
        writer.writerow(['ID', 'Area', 'Solidity', 'Max Width', 'Max Height', 'Curvature'])
        for result in results:
            writer.writerow([result.id, result.area, result.solidity, result.max_height, result.max_width, result.curvature])

    # Excel
    excel_path = join(output_directory, f"{options.input_stem}.results.xlsx")
    print(f"Writing Excel file: {excel_path}")
    wb = load_workbook(excel_path) if isfile(excel_path) else Workbook()
    sheet = wb.active
    sheet.cell(row=1, column=1).value = 'filename'
    sheet.cell(row=1, column=2).value = 'leaf_area'
    sheet.cell(row=1, column=3).value = 'solidity'
    sheet.cell(row=1, column=4).value = 'max_width'
    sheet.cell(row=1, column=5).value = 'max_height'
    sheet.cell(row=1, column=6).value = 'curvature'
    for result in results:
        sheet.append([result.id, result.area, result.solidity, result.max_height, result.max_width, result.curvature])
    wb.save(excel_path)


if __name__ == '__main__':
    cli()
