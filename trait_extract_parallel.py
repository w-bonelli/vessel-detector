"""
Name: trait_extract_parallel.py

Version: 1.0

Summary: Extract plant traits (leaf area, width, height, solidity, curvature) by parallel processing

Author: Suxing Liu

Author-email: suxingliu@gmail.com

Created: 2018-09-29

USAGE:

time python3 trait_extract_parallel.py -i /input/directory -o /output/directory -ft jpg
"""

import argparse
import glob
import os
import warnings
from collections import Counter
from os.path import join, dirname
from pathlib import Path

import cv2
import imutils
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import Workbook
from openpyxl import load_workbook
from scipy import ndimage
from scipy.interpolate import interp1d
from scipy.spatial import distance as dist
from skimage import img_as_float, img_as_ubyte, img_as_bool
from skimage.color import rgb2lab, deltaE_cie76
from skimage.feature import peak_local_max
from skimage.morphology import watershed, medial_axis
from sklearn.cluster import KMeans
from sklearn.cluster import MiniBatchKMeans

from tools import utils
from tools.curvature import ComputeCurvature

warnings.filterwarnings("ignore")

import multiprocessing
from multiprocessing import Pool
from contextlib import closing

MBFACTOR = float(1 << 20)


def color_cluster_seg(image, args_colorspace, args_channels, args_num_clusters):
    # Change image color space, if necessary.
    colorSpace = args_colorspace.lower()

    if colorSpace == 'hsv':
        image = cv2.cvtColor(image, cv2.COLOR_BGR2HSV)

    elif colorSpace == 'ycrcb' or colorSpace == 'ycc':
        image = cv2.cvtColor(image, cv2.COLOR_BGR2YCrCb)

    elif colorSpace == 'lab':
        image = cv2.cvtColor(image, cv2.COLOR_BGR2LAB)

    else:
        colorSpace = 'bgr'  # set for file naming purposes

    # Keep only the selected channels for K-means clustering.
    if args_channels != 'all':
        channels = cv2.split(image)
        channelIndices = []
        for char in args_channels:
            channelIndices.append(int(char))
        image = image[:, :, channelIndices]
        if len(image.shape) == 2:
            image.reshape(image.shape[0], image.shape[1], 1)

    (width, height, n_channel) = image.shape

    # Flatten the 2D image array into an MxN feature vector, where M is the number of pixels and N is the dimension (number of channels).
    reshaped = image.reshape(image.shape[0] * image.shape[1], image.shape[2])

    # Perform K-means clustering.
    if args_num_clusters < 2:
        print('Warning: num-clusters < 2 invalid. Using num-clusters = 2')

    # define number of cluster
    numClusters = max(2, args_num_clusters)

    # clustering method
    kmeans = KMeans(n_clusters=numClusters, n_init=40, max_iter=500).fit(reshaped)

    # get lables 
    pred_label = kmeans.labels_

    # Reshape result back into a 2D array, where each element represents the corresponding pixel's cluster index (0 to K - 1).
    clustering = np.reshape(np.array(pred_label, dtype=np.uint8), (image.shape[0], image.shape[1]))

    # Sort the cluster labels in order of the frequency with which they occur.
    sortedLabels = sorted([n for n in range(numClusters)], key=lambda x: -np.sum(clustering == x))

    # Initialize K-means grayscale image; set pixel colors based on clustering.
    kmeansImage = np.zeros(image.shape[:2], dtype=np.uint8)
    for i, label in enumerate(sortedLabels):
        kmeansImage[clustering == label] = int(255 / (numClusters - 1)) * i

    ret, thresh = cv2.threshold(kmeansImage, 0, 255, cv2.THRESH_BINARY | cv2.THRESH_OTSU)

    # return thresh

    nb_components, output, stats, centroids = cv2.connectedComponentsWithStats(thresh, connectivity=8)

    sizes = stats[1:, -1]

    nb_components = nb_components - 1

    min_size = 150

    img_thresh = np.zeros([width, height], dtype=np.uint8)

    # for every component in the image, you keep it only if it's above min_size
    for i in range(0, nb_components):
        if sizes[i] >= min_size:
            img_thresh[output == i + 1] = 255

    # from skimage import img_as_ubyte

    # img_thresh = img_as_ubyte(img_thresh)

    # print("img_thresh.dtype")
    # print(img_thresh.dtype)

    return img_thresh


def medial_axis_image(thresh):
    # convert an image from OpenCV to skimage
    thresh_sk = img_as_float(thresh)

    image_bw = img_as_bool((thresh_sk))

    image_medial_axis = medial_axis(image_bw)

    return image_medial_axis


def watershed_seg(orig, thresh, min_distance_value):
    # compute the exact Euclidean distance from every binary
    # pixel to the nearest zero pixel, then find peaks in this
    # distance map
    D = ndimage.distance_transform_edt(thresh)

    localMax = peak_local_max(D, indices=False, min_distance=min_distance_value, labels=thresh)

    # perform a connected component analysis on the local peaks,
    # using 8-connectivity, then appy the Watershed algorithm
    markers = ndimage.label(localMax, structure=np.ones((3, 3)))[0]

    labels = watershed(-D, markers, mask=thresh)

    print(f"{len(np.unique(labels)) - 1} unique segments found")

    return labels


def comp_external_contour(orig, thresh):
    # find contours and get the external one
    contours, hier = cv2.findContours(thresh, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

    img_height, img_width, img_channels = orig.shape

    index = 1

    for c in contours:

        # get the bounding rect
        x, y, w, h = cv2.boundingRect(c)

        if w > img_width * 0.1 and h > img_height * 0.1:

            trait_img = cv2.drawContours(orig, contours, -1, (255, 255, 0), 1)

            # draw a green rectangle to visualize the bounding rect
            roi = orig[y:y + h, x:x + w]

            print(f"ROI {index} detected ...")
            # result_file = (save_path +  str(index) + file_extension)
            # cv2.imwrite(result_file, roi)

            trait_img = cv2.rectangle(orig, (x, y), (x + w, y + h), (255, 255, 0), 3)

            index += 1

            '''
            #get the min area rect
            rect = cv2.minAreaRect(c)
            box = cv2.boxPoints(rect)
            # convert all coordinates floating point values to int
            box = np.int0(box)
            #draw a red 'nghien' rectangle
            trait_img = cv2.drawContours(orig, [box], 0, (0, 0, 255))
            '''
            # get convex hull
            hull = cv2.convexHull(c)
            # draw it in red color
            trait_img = cv2.drawContours(orig, [hull], -1, (0, 0, 255), 3)

            '''
            # calculate epsilon base on contour's perimeter
            # contour's perimeter is returned by cv2.arcLength
            epsilon = 0.01 * cv2.arcLength(c, True)
            # get approx polygons
            approx = cv2.approxPolyDP(c, epsilon, True)
            # draw approx polygons
            trait_img = cv2.drawContours(orig, [approx], -1, (0, 255, 0), 1)
         
            # hull is convex shape as a polygon
            hull = cv2.convexHull(c)
            trait_img = cv2.drawContours(orig, [hull], -1, (0, 0, 255))
            '''

            '''
            #get the min enclosing circle
            (x, y), radius = cv2.minEnclosingCircle(c)
            # convert all values to int
            center = (int(x), int(y))
            radius = int(radius)
            # and draw the circle in blue
            trait_img = cv2.circle(orig, center, radius, (255, 0, 0), 2)
            '''

            area = cv2.contourArea(c)
            print("Area: {0:.2f} ".format(area))

            hull = cv2.convexHull(c)
            hull_area = cv2.contourArea(hull)
            solidity = float(area) / hull_area
            print("Solidity: {0:.2f} ".format(solidity))

            extLeft = tuple(c[c[:, :, 0].argmin()][0])
            extRight = tuple(c[c[:, :, 0].argmax()][0])
            extTop = tuple(c[c[:, :, 1].argmin()][0])
            extBot = tuple(c[c[:, :, 1].argmax()][0])

            trait_img = cv2.circle(orig, extLeft, 3, (255, 0, 0), -1)
            trait_img = cv2.circle(orig, extRight, 3, (255, 0, 0), -1)
            trait_img = cv2.circle(orig, extTop, 3, (255, 0, 0), -1)
            trait_img = cv2.circle(orig, extBot, 3, (255, 0, 0), -1)

            max_width = dist.euclidean(extLeft, extRight)
            max_height = dist.euclidean(extTop, extBot)

            if max_width > max_height:
                trait_img = cv2.line(orig, extLeft, extRight, (0, 255, 0), 2)
            else:
                trait_img = cv2.line(orig, extTop, extBot, (0, 255, 0), 2)

            print("Width, Height: {0:.2f}, {1:.2f} ".format(w, h))

            return trait_img, area, solidity, w, h
        return None, None, None, w, h
    return None, None, None, None, None



def compute_curv(orig, labels):
    gray = cv2.cvtColor(orig, cv2.COLOR_BGR2GRAY)

    curv_sum = 0.0
    count = 0
    # curvature computation
    # loop over the unique labels returned by the Watershed algorithm
    for index, label in enumerate(np.unique(labels), start=1):
        # if the label is zero, we are examining the 'background'
        # so simply ignore it
        if label == 0:
            continue

        # otherwise, allocate memory for the label region and draw
        # it on the mask
        mask = np.zeros(gray.shape, dtype="uint8")
        mask[labels == label] = 255

        # detect contours in the mask and grab the largest one
        # cnts = cv2.findContours(mask.copy(), cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)[-2]
        contours, hierarchy = cv2.findContours(mask.copy(), cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)
        c = max(contours, key=cv2.contourArea)

        # draw a circle enclosing the object
        ((x, y), r) = cv2.minEnclosingCircle(c)
        label_trait = cv2.circle(orig, (int(x), int(y)), 3, (0, 255, 0), 2)
        label_trait = cv2.putText(orig, "#{}".format(label), (int(x) - 10, int(y)), cv2.FONT_HERSHEY_SIMPLEX, 0.6,
                                  (0, 0, 255), 2)
        # cv2.putText(orig, "#{}".format(curvature), (int(x) - 10, int(y)), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 0, 255), 2)

        if len(c) >= 5:
            label_trait = cv2.drawContours(orig, [c], -1, (255, 0, 0), 2)
            ellipse = cv2.fitEllipse(c)
            label_trait = cv2.ellipse(orig, ellipse, (0, 255, 0), 2)

            c_np = np.vstack(c).squeeze()
            count += 1

            x = c_np[:, 0]
            y = c_np[:, 1]

            comp_curv = ComputeCurvature(x, y)
            curvature = comp_curv.fit(x, y)

            curv_sum = curv_sum + curvature

        else:
            # optional to "delete" the small contours
            label_trait = cv2.drawContours(orig, [c], -1, (0, 0, 255), 2)
            print("Not enough points to fit ellipse")

    print('Average curvature: {0:.2f}'.format(curv_sum / count))

    return curv_sum / count, label_trait


def RGB2HEX(color):
    return "#{:02x}{:02x}{:02x}".format(int(color[0]), int(color[1]), int(color[2]))


def color_quantization(image, mask, save_path, num_clusters):
    # grab image width and height
    (h, w) = image.shape[:2]

    # change the color storage order
    image = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)

    # apply the mask to get the segmentation of plant
    masked_image = cv2.bitwise_and(image, image, mask=mask)

    # reshape the image to be a list of pixels
    pixels = masked_image.reshape((masked_image.shape[0] * masked_image.shape[1], 3))

    ############################################################
    # Clustering process
    ###############################################################
    # cluster the pixel intensities
    clt = MiniBatchKMeans(n_clusters=num_clusters)
    # clt = KMeans(n_clusters = args["clusters"])
    clt.fit(pixels)

    # assign labels to each cluster
    labels = clt.fit_predict(pixels)

    # obtain the quantized clusters using each label
    quant = clt.cluster_centers_.astype("uint8")[labels]

    # reshape the feature vectors to images
    quant = quant.reshape((h, w, 3))
    image_rec = pixels.reshape((h, w, 3))

    # convert from L*a*b* to RGB
    quant = cv2.cvtColor(quant, cv2.COLOR_RGB2BGR)
    image_rec = cv2.cvtColor(image_rec, cv2.COLOR_RGB2BGR)

    # display the images and wait for a keypress
    # cv2.imshow("image", np.hstack([image_rec, quant]))
    # cv2.waitKey(0)

    # define result path for labeled images
    result_img_path = save_path + 'cluster_out.png'

    # save color_quantization results
    cv2.imwrite(result_img_path, quant)

    # Get colors and analze them from masked image
    counts = Counter(labels)
    # sort to ensure correct color percentage
    counts = dict(sorted(counts.items()))

    center_colors = clt.cluster_centers_

    # print(type(center_colors))

    # We get ordered colors by iterating through the keys
    ordered_colors = [center_colors[i] for i in counts.keys()]
    hex_colors = [RGB2HEX(ordered_colors[i]) for i in counts.keys()]
    rgb_colors = [ordered_colors[i] for i in counts.keys()]

    # print(hex_colors)

    index_bkg = [index for index in range(len(hex_colors)) if hex_colors[index] == '#000000']

    # print(index_bkg[0])

    # print(counts)
    # remove background color
    del hex_colors[index_bkg[0]]
    del rgb_colors[index_bkg[0]]

    # Using dictionary comprehension to find list 
    # keys having value . 
    delete = [key for key in counts if key == index_bkg[0]]

    # delete the key 
    for key in delete: del counts[key]

    fig = plt.figure(figsize=(6, 6))
    plt.pie(counts.values(), labels=hex_colors, colors=hex_colors)

    # define result path for labeled images
    result_img_path = save_path + 'pie_color.png'
    plt.savefig(result_img_path)

    # build a histogram of clusters and then create a figure representing the number of pixels labeled to each color
    hist = utils.centroid_histogram(clt)

    # remove the background color cluster
    clt.cluster_centers_ = np.delete(clt.cluster_centers_, index_bkg[0], axis=0)

    # build a histogram of clusters using center lables
    numLabels = utils.plot_centroid_histogram(save_path, clt)

    # create a figure representing the distribution of each color
    bar = utils.plot_colors(hist, clt.cluster_centers_)

    # save a figure of color bar
    utils.plot_color_bar(save_path, bar)

    return rgb_colors


def get_cmap(n, name='hsv'):
    '''Returns a function that maps each index in 0, 1, ..., n-1 to a distinct 
    RGB color; the keyword argument name must be a standard mpl colormap name.'''
    return plt.cm.get_cmap(name, n)


def color_region(image, mask, output_dir, num_clusters):
    # read the image
    # grab image width and height
    (h, w) = image.shape[:2]

    # apply the mask to get the segmentation of plant
    masked_image_ori = cv2.bitwise_and(image, image, mask=mask)

    # define result path for labeled images
    drawn_contours_path = output_dir + '_masked.png'
    cv2.imwrite(drawn_contours_path, masked_image_ori)

    # convert to RGB
    image_RGB = cv2.cvtColor(masked_image_ori, cv2.COLOR_BGR2RGB)

    # reshape the image to a 2D array of pixels and 3 color values (RGB)
    pixel_values = image_RGB.reshape((-1, 3))

    # convert to float
    pixel_values = np.float32(pixel_values)

    # define stopping criteria
    criteria = (cv2.TERM_CRITERIA_EPS + cv2.TERM_CRITERIA_MAX_ITER, 100, 0.2)

    # number of clusters (K)
    # num_clusters = 5
    compactness, labels, (centers) = cv2.kmeans(pixel_values, num_clusters, None, criteria, 10,
                                                cv2.KMEANS_RANDOM_CENTERS)

    # convert back to 8 bit values
    centers = np.uint8(centers)

    # flatten the labels array
    labels_flat = labels.flatten()

    # convert all pixels to the color of the centroids
    segmented_image = centers[labels_flat]

    # reshape back to the original image dimension
    segmented_image = segmented_image.reshape(image_RGB.shape)

    segmented_image_BRG = cv2.cvtColor(segmented_image, cv2.COLOR_RGB2BGR)
    # define result path for labeled images
    drawn_contours_path = output_dir + '_clustered.png'
    cv2.imwrite(drawn_contours_path, segmented_image_BRG)

    '''
    fig = plt.figure()
    ax = Axes3D(fig)        
    for label, pix in zip(labels, segmented_image):
        ax.scatter(pix[0], pix[1], pix[2], color = (centers))
            
    result_file = (save_path + base_name + 'color_cluster_distributation.png')
    plt.savefig(result_file)
    '''
    # Show only one chosen cluster
    # masked_image = np.copy(image)
    masked_image = np.zeros_like(image_RGB)

    # convert to the shape of a vector of pixel values
    masked_image = masked_image.reshape((-1, 3))
    # color (i.e cluster) to render
    # cluster = 2

    cmap = get_cmap(num_clusters + 1)

    # clrs = sns.color_palette('husl', n_colors = num_clusters)  # a list of RGB tuples

    color_conversion = interp1d([0, 1], [0, 255])

    for cluster in range(num_clusters):

        print("Processing cluster {0}...".format(cluster))
        # print(clrs[cluster])
        # print(color_conversion(clrs[cluster]))

        masked_image[labels_flat == cluster] = centers[cluster]

        # print(centers[cluster])

        # convert back to original shape
        masked_image_rp = masked_image.reshape(image_RGB.shape)

        # masked_image_BRG = cv2.cvtColor(masked_image, cv2.COLOR_RGB2BGR)
        # cv2.imwrite('maksed.png', masked_image_BRG)

        gray = cv2.cvtColor(masked_image_rp, cv2.COLOR_BGR2GRAY)

        # threshold the image, then perform a series of erosions +
        # dilations to remove any small regions of noise
        thresh = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY)[1]

        contours = cv2.findContours(thresh.copy(), cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        contours = imutils.grab_contours(contours)
        # c = max(cnts, key=cv2.contourArea)

        '''
        # compute the center of the contour area and draw a circle representing the center
        M = cv2.moments(c)
        cX = int(M["m10"] / M["m00"])
        cY = int(M["m01"] / M["m00"])
        # draw the countour number on the image
        result = cv2.putText(masked_image_rp, "#{}".format(cluster + 1), (cX - 20, cY), cv2.FONT_HERSHEY_SIMPLEX, 1.0, (255, 255, 255), 2)
        '''

        if not contours:
            print("findContours is empty")
        else:
            # loop over the (unsorted) contours and draw them
            drawn_contours = None
            for (i, c) in enumerate(contours):
                drawn_contours = cv2.drawContours(masked_image_rp, c, -1, color_conversion(np.random.random(3)), 2)
                # result = cv2.drawContours(masked_image_rp, c, -1, color_conversion(clrs[cluster]), 2)

            if drawn_contours is None:
                return

            drawn_contours[drawn_contours == 0] = 255

            drawn_contours_colors = cv2.cvtColor(drawn_contours, cv2.COLOR_RGB2BGR)
            drawn_contours_path = output_dir + '_contours_' + str(cluster) + '.png'
            cv2.imwrite(drawn_contours_path, drawn_contours_colors)

    counts = Counter(labels_flat)
    # sort to ensure correct color percentage
    counts = dict(sorted(counts.items()))

    center_colors = centers

    # We get ordered colors by iterating through the keys
    ordered_colors = [center_colors[i] for i in counts.keys()]
    hex_colors = [RGB2HEX(ordered_colors[i]) for i in counts.keys()]
    rgb_colors = [ordered_colors[i] for i in counts.keys()]

    # print(hex_colors)

    index_bkg = [index for index in range(len(hex_colors)) if hex_colors[index] == '#000000']

    # print(index_bkg[0])

    # print(counts)
    # remove background color
    del hex_colors[index_bkg[0]]
    del rgb_colors[index_bkg[0]]

    # Using dictionary comprehension to find list 
    # keys having value . 
    delete = [key for key in counts if key == index_bkg[0]]

    # delete the key 
    for key in delete: del counts[key]

    fig = plt.figure(figsize=(6, 6))
    plt.pie(counts.values(), labels=hex_colors, colors=hex_colors)

    # define result path for labeled images
    drawn_contours_path = output_dir + '_pie_color.png'
    plt.savefig(drawn_contours_path)

    return rgb_colors


def extract_traits(image_path):
    print(image_path)
    image_abs_path = os.path.abspath(image_path)
    image_file_name, img_file_ext = os.path.splitext(image_abs_path)
    image_file_size = os.path.getsize(image_path) / MBFACTOR
    image_file = os.path.splitext(os.path.basename(image_file_name))[0]

    output_dir = join(dirname(dirname(image_path)), 'output')
    Path(output_dir).mkdir(exist_ok=True)

    print(f"Extracting traits for image '{image_file}' to '{output_dir}'... Segmenting image using automatic color clustering..." + f" File size is large: {str(image_file_size)} MB. This may take some time." if image_file_size > 5.0 else '')

    image = cv2.imread(image_path)

    # make backup image
    orig = image.copy()

    args_colorspace = args['color_space']
    args_channels = args['channels']
    args_num_clusters = args['num_clusters']

    # color clustering based plant object segmentation
    thresh = color_cluster_seg(orig, args_colorspace, args_channels, args_num_clusters)

    # save segmentation result
    seg = join(output_dir, f"{image_file}_seg{img_file_ext}")
    cv2.imwrite(seg, thresh)

    num_clusters = 5
    rgb_colors = color_region(orig, thresh, join(output_dir, image_file), num_clusters)

    print("Color difference:")

    selected_color = rgb2lab(np.uint8(np.asarray([[rgb_colors[0]]])))

    for index, value in enumerate(rgb_colors):
        # print(index, value)
        curr_color = rgb2lab(np.uint8(np.asarray([[value]])))
        diff = deltaE_cie76(selected_color, curr_color)
        print(index, value, diff)

    # accquire medial axis of segmentation mask
    image_medial_axis = medial_axis_image(thresh)

    # save medial axis result
    maxis = join(output_dir, f"{image_file}_medial_axis{img_file_ext}")
    cv2.imwrite(maxis, img_as_ubyte(image_medial_axis))

    min_distance_value = 5
    # watershed based leaf area segmentaiton
    labels = watershed_seg(orig, thresh, min_distance_value)

    # save watershed result label image
    # Map component labels to hue val
    label_hue = np.uint8(128 * labels / np.max(labels))
    # label_hue[labels == largest_label] = np.uint8(15)
    blank_ch = 255 * np.ones_like(label_hue)
    labeled_img = cv2.merge([label_hue, blank_ch, blank_ch])

    # cvt to BGR for display
    labeled_img = cv2.cvtColor(labeled_img, cv2.COLOR_HSV2BGR)

    # set background label to black
    labeled_img[label_hue == 0] = 0
    lbl = join(output_dir, f"{image_file}_label{img_file_ext}")
    cv2.imwrite(lbl, labeled_img)

    # find curvature
    (avg_curv, label_trait) = compute_curv(orig, labels)
    curv = join(output_dir, f"{image_file}_curv{img_file_ext}")
    cv2.imwrite(curv, label_trait)

    # find external contour
    (trait_img, area, solidity, max_width, max_height) = comp_external_contour(image.copy(), thresh)
    excont = join(output_dir, f"{image_file}_excontour{img_file_ext}")
    if trait_img is not None:
        cv2.imwrite(excont, trait_img)

    return image_file_name, area, solidity, max_width, max_height, avg_curv


if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument("-i", "--input-path", required=True, help="path to input directory (containing image files)")
    parser.add_argument("-o", "--output-path", required=False, help="path to directory to write output files")
    parser.add_argument("-ft", "--filetype", required=True, help="Image filetype")
    parser.add_argument('-s', '--color-space', type=str, default='lab',
                        help='Color space to use: BGR (default), HSV, Lab, YCrCb (YCC)')
    parser.add_argument('-c', '--channels', type=str, default='1',
                        help='Channel indices to use for clustering, where 0 is the first channel,'
                         + ' 1 is the second channel, etc. E.g., if BGR color space is used, "02" '
                         + 'selects channels B and R. (default "all")')
    parser.add_argument('-n', '--num-clusters', type=int, default=2,
                        help='Number of clusters for K-means clustering (default 3, min 2).')

    args = vars(parser.parse_args())
    input_dir = args["input-path"]
    output_dir = args["output-path"] if 'output-path' in args else None
    input_images = sorted(glob.glob(f"{input_dir}/*.{args['filetype']}"))
    cpus = multiprocessing.cpu_count()

    print(f"Using {int(cpus)} cores to process {len(input_images)} images...")

    with closing(Pool(processes=cpus)) as pool:
        result = pool.map(extract_traits, input_images)
        pool.terminate()

    # if output dir provided, create it (if needed). otherwise use current directory
    if output_dir is not None:
        Path(output_dir).mkdir(exist_ok=True)
    else:
        output_dir = ''

    trait_file = join(output_dir, 'trait.xlsx')
    print(f"Writing trait file '{trait_file}'...")

    if os.path.isfile(trait_file):
        wb = load_workbook(trait_file)
    else:
        wb = Workbook()

    sheet = wb.active
    sheet.cell(row=1, column=1).value = 'filename'
    sheet.cell(row=1, column=2).value = 'leaf_area'
    sheet.cell(row=1, column=3).value = 'solidity'
    sheet.cell(row=1, column=4).value = 'max_width'
    sheet.cell(row=1, column=5).value = 'max_height'
    sheet.cell(row=1, column=6).value = 'curvature'
    for row in result:
        sheet.append(row)

    wb.save(trait_file)
